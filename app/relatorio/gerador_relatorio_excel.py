"""
Gerador de Relatórios Excel
OrionForgeNexus (oFn)

Gera relatórios Excel para estruturas de banco de dados e resultados de queries.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from app.schemas.query_select_upAndInsert_schema import QueryResultType
from app.ultils.logger import log_message


# ============================================================
# 🎨 CONSTANTES DE ESTILO
# ============================================================

class ExcelStyles:
    """Definições de estilos para Excel"""
    
    COLOR_PRIMARY = "1e3a8a"
    COLOR_SECONDARY = "0ea5e9"
    COLOR_SUCCESS = "059669"
    COLOR_INFO = "1d4ed8"
    COLOR_TEXT_PRIMARY = "475569"
    COLOR_TEXT_SECONDARY = "64748b"
    COLOR_WHITE = "FFFFFF"
    COLOR_BORDER = "cbd5e1"
    COLOR_ALTERNATE_ROW = "f8fafc"
    COLOR_LIGHT_BLUE = "f0f9ff"
    
    FONT_TITLE = Font(bold=True, size=16, color=COLOR_PRIMARY)
    FONT_INFO = Font(size=10, color=COLOR_TEXT_SECONDARY)
    FONT_HEADER_SECONDARY = Font(bold=True, color=COLOR_WHITE, size=10)
    FONT_DATA = Font(size=10, color=COLOR_TEXT_PRIMARY)
    FONT_SECTION = Font(bold=True, size=12, color=COLOR_PRIMARY)
    
    FILL_HEADER_INFO = PatternFill(start_color=COLOR_INFO, end_color=COLOR_INFO, fill_type="solid")
    FILL_ALTERNATE = PatternFill(start_color=COLOR_ALTERNATE_ROW, end_color=COLOR_ALTERNATE_ROW, fill_type="solid")
    FILL_LIGHT_BLUE = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
    
    BORDER_THIN = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )
    
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ============================================================
# 🏗️ CLASSE PRINCIPAL
# ============================================================

class ExcelReportGenerator:
    """Gerador de relatórios Excel com template OrionForgeNexus"""
    
    DEFAULT_COMPANY_NAME = "OrionForgeNexus"
    DEFAULT_COMPANY_ABBR = "oFn"
    LOGO_SIZE = (80, 80)
    MAX_COLUMN_WIDTH = 60
    MIN_COLUMN_WIDTH = 12
    
    def __init__(self, logo_path: Optional[str] = None, company_name: str = DEFAULT_COMPANY_NAME, company_abbr: str = DEFAULT_COMPANY_ABBR):
        self.logo_path = self._validate_logo_path(logo_path)
        self.company_name = company_name
        self.company_abbr = company_abbr
    
    @staticmethod
    def _validate_logo_path(logo_path: Optional[str]) -> Optional[Path]:
        if not logo_path: return None
        path = Path(logo_path)
        return path if path.is_file() else None
    
    def _add_header(self, ws: Worksheet, titulo: str, total_cols: int = 8) -> int:
        """Adiciona o cabeçalho padrão do relatório."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(row=1, column=1, value=titulo).font = ExcelStyles.FONT_TITLE
        ws.cell(row=1, column=1).alignment = ExcelStyles.ALIGN_CENTER

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        date_text = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws.cell(row=2, column=1, value=f"Gerado em {date_text} - {self.company_name} ({self.company_abbr})").font = ExcelStyles.FONT_INFO
        ws.cell(row=2, column=1).alignment = ExcelStyles.ALIGN_CENTER

        if self.logo_path:
            try:
                img = XLImage(str(self.logo_path))
                img.width, img.height = self.LOGO_SIZE
                ws.add_image(img, "A1")
            except Exception as e:
                log_message(f"⚠️ Falha ao carregar logo: {e}", "warning")

        return 4

    def _auto_adjust_columns(self, ws: Worksheet, total_cols: int) -> None:
        """Ajusta automaticamente a largura das colunas."""
        for col_idx in range(1, total_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max((len(str(cell.value)) for cell in ws[col_letter] if cell.value), default=0)
            ws.column_dimensions[col_letter].width = max(self.MIN_COLUMN_WIDTH, min(max_length + 2, self.MAX_COLUMN_WIDTH))

    # ============================================================
    # 📋 RELATÓRIO DE QUERY
    # ============================================================
    
    def generate_query_report(self, filename: str, query_result: QueryResultType) -> None:
        """Gera relatório Excel de resultado de query."""
        log_message("📋 Gerando relatório Excel de query...")
        wb = Workbook()
        
        ws_info = wb.active
        ws_info.title = "Informações"
        self._add_query_info_sheet(ws_info, query_result)
        
        ws_results = wb.create_sheet("Resultados")
        self._add_query_results_sheet(ws_results, query_result)
        
        wb.save(filename)
        log_message(f"✅ Relatório Excel de query gerado: {filename}")
    
    def _add_query_info_sheet(self, ws: Worksheet, query_result: QueryResultType) -> None:
        current_row = self._add_header(ws, "Relatório de Execução de Query SQL", 4)
        
        ws[f"A{current_row}"] = "Status da Execução"
        ws[f"A{current_row}"].font = ExcelStyles.FONT_SECTION
        current_row += 1
        
        status_data = [
            ["Status", "✅ Sucesso" if query_result.success else "❌ Falhou"],
            ["Duração (ms)", str(query_result.duration_ms)],
            ["Total de Resultados", str(query_result.totalResults or "—")],
            ["Prévia Carregada", str(len(query_result.preview))],
            ["Query SQL", query_result.query]
        ]
        
        for row_data in status_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = ExcelStyles.BORDER_THIN
                cell.alignment = ExcelStyles.ALIGN_LEFT
                
                if col_idx == 1:
                    cell.font = Font(bold=True, size=10)
                    cell.fill = ExcelStyles.FILL_LIGHT_BLUE
                else:
                    cell.fill = ExcelStyles.FILL_ALTERNATE
            current_row += 1
            
        self._auto_adjust_columns(ws, 2)

    def _add_query_results_sheet(self, ws: Worksheet, query_result: QueryResultType) -> None:
        if not query_result.columns or not query_result.preview:
            ws["A1"] = "Nenhum dado retornado pela query."
            ws["A1"].font = ExcelStyles.FONT_INFO
            return

        MAX_COLS = 15 # Aumentei um pouco para Excel, já que cabe mais que num PDF
        columns = query_result.columns
        truncated = len(columns) > MAX_COLS
        
        # Limpa nomes ("main.bank_full.age" -> "age")
        visible_columns = [c.split('.')[-1] for c in columns[:MAX_COLS]]
        if truncated: visible_columns.append("...")

        current_row = 1
        ws.merge_cells(f"A{current_row}:{get_column_letter(len(visible_columns))}{current_row}")
        ws[f"A{current_row}"] = "📄 Resultados da Query"
        ws[f"A{current_row}"].font = ExcelStyles.FONT_SECTION
        current_row += 2

        # Cabeçalhos
        for col_idx, col_name in enumerate(visible_columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill, cell.font, cell.alignment, cell.border = ExcelStyles.FILL_HEADER_INFO, ExcelStyles.FONT_HEADER_SECONDARY, ExcelStyles.ALIGN_CENTER, ExcelStyles.BORDER_THIN
        current_row += 1

        # Dados
        for record in query_result.preview:
            row_data = [str(record.get(col, "")) for col in columns[:MAX_COLS]]
            if truncated: row_data.append("...")
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font, cell.alignment, cell.border = ExcelStyles.FONT_DATA, ExcelStyles.ALIGN_LEFT, ExcelStyles.BORDER_THIN
                if current_row % 2 == 0: cell.fill = ExcelStyles.FILL_ALTERNATE
            current_row += 1

        if truncated:
            current_row += 1
            ws.merge_cells(f"A{current_row}:{get_column_letter(len(visible_columns))}{current_row}")
            cell = ws[f"A{current_row}"]
            cell.value = f"⚠️ Exibindo apenas {MAX_COLS} de {len(columns)} colunas."
            cell.font, cell.alignment = ExcelStyles.FONT_INFO, ExcelStyles.ALIGN_LEFT

        self._auto_adjust_columns(ws, len(visible_columns))

    # ============================================================
    # 📊 RELATÓRIO DE ESTRUTURA DE BANCO
    # ============================================================
    # Mantive a sua lógica de estrutura intocável, apenas converti para o novo design de classes limpas.
    def generate_structure_report(self, filename: str, metadata_list: List[Dict[str, Any]]) -> None:
        if not metadata_list: raise ValueError("Lista de metadados não pode estar vazia")
        log_message("📊 Gerando relatório Excel de estrutura...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Estrutura do Banco"
        current_row = self._add_header(ws, "Relatório de Estrutura do Banco de Dados", 8)
        
        for idx, metadata in enumerate(metadata_list, 1):
            current_row = self._add_table_structure(ws, metadata, current_row, idx) + 2
            
        self._auto_adjust_columns(ws, 8)
        wb.save(filename)
        log_message(f"✅ Relatório Excel de estrutura gerado: {filename}")
    
    def _add_table_structure(self, ws: Worksheet, metadata: Dict[str, Any], start_row: int, table_number: int) -> int:
        ws.merge_cells(f"A{start_row}:H{start_row}")
        cell = ws[f"A{start_row}"]
        cell.value = f"Tabela {table_number}: {metadata.get('schema_name', '')}.{metadata.get('table_name', '')}"
        cell.font, cell.alignment = ExcelStyles.FONT_SECTION, ExcelStyles.ALIGN_LEFT
        start_row += 1
        
        info_data = [
            ["Schema", metadata.get("schema_name", "—")],
            ["Tabela", metadata.get("table_name", "—")],
            ["Total de Colunas", metadata.get("total_colunas", 0)],
            ["Executado em", metadata.get("executado_em", "—")],
        ]
        
        for row_data in info_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=value)
                cell.border, cell.alignment = ExcelStyles.BORDER_THIN, ExcelStyles.ALIGN_LEFT
                if col_idx == 1: cell.font, cell.fill = Font(bold=True, size=10), ExcelStyles.FILL_LIGHT_BLUE
                else: cell.fill = ExcelStyles.FILL_ALTERNATE
            start_row += 1
        
        start_row += 1
        for col_idx, col_name in enumerate(["Nome", "Tipo", "Nullable", "PK", "FK", "Unique", "Default", "Comentário"], 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.fill, cell.font, cell.alignment, cell.border = ExcelStyles.FILL_HEADER_INFO, ExcelStyles.FONT_HEADER_SECONDARY, ExcelStyles.ALIGN_CENTER, ExcelStyles.BORDER_THIN
        start_row += 1
        
        for coluna in metadata.get("colunas", []):
            row_data = [
                coluna.get("nome", ""), coluna.get("tipo", ""), "Sim" if coluna.get("is_nullable") else "Não",
                "✓" if coluna.get("is_primary_key") else "", "✓" if coluna.get("is_foreign_key") else "",
                "✓" if coluna.get("is_unique") else "", coluna.get("default", ""), coluna.get("comentario", ""),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=value)
                cell.border, cell.font = ExcelStyles.BORDER_THIN, ExcelStyles.FONT_DATA
                cell.alignment = ExcelStyles.ALIGN_LEFT if col_idx in [1, 2, 8] else ExcelStyles.ALIGN_CENTER
                if start_row % 2 == 0: cell.fill = ExcelStyles.FILL_ALTERNATE
            start_row += 1
        
        return start_row